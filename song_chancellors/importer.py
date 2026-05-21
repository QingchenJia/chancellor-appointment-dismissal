from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from .db import connect, create_schema, rebuild_database
from .parsing import classify_event_type, normalize_month, split_person_name


def import_workbook(workbook_path: str | Path, db_path: str | Path, rebuild: bool = False) -> dict[str, Any]:
    source = Path(workbook_path)
    conn = rebuild_database(db_path) if rebuild else connect(db_path)
    create_schema(conn)

    wb = openpyxl.load_workbook(source, data_only=True)
    ws = wb.worksheets[0]

    persons: dict[int, int] = {}
    for col in range(5, ws.max_column + 1):
        raw_name = ws.cell(2, col).value
        if not raw_name:
            continue
        source_column = get_column_letter(col)
        parsed = split_person_name(str(raw_name), source_column)
        cur = conn.execute(
            """
            insert into persons (canonical_name, raw_name, aliases, source_column)
            values (?, ?, ?, ?)
            on conflict(source_column) do update set
                canonical_name = excluded.canonical_name,
                raw_name = excluded.raw_name,
                aliases = excluded.aliases
            returning id
            """,
            (parsed.canonical_name, parsed.raw_name, parsed.aliases, parsed.source_column),
        )
        persons[col] = int(cur.fetchone()["id"])

    current_year = None
    current_emperor = None
    current_era = None
    time_points: dict[int, int] = {}
    record_count = 0
    comment_count = 0
    warning_count = 0

    for row in range(3, ws.max_row + 1):
        year_value = ws.cell(row, 1).value
        emperor_value = ws.cell(row, 2).value
        era_value = ws.cell(row, 3).value
        month_label = ws.cell(row, 4).value
        if year_value not in (None, ""):
            current_year = int(year_value)
        if emperor_value not in (None, ""):
            current_emperor = str(emperor_value)
        if era_value not in (None, ""):
            current_era = str(era_value)
        if month_label in (None, ""):
            continue

        month_index = normalize_month(str(month_label))
        cur = conn.execute(
            """
            insert into time_points (gregorian_year, month_index, month_label, emperor, era_name, source_row)
            values (?, ?, ?, ?, ?, ?)
            on conflict(source_row) do update set
                gregorian_year = excluded.gregorian_year,
                month_index = excluded.month_index,
                month_label = excluded.month_label,
                emperor = excluded.emperor,
                era_name = excluded.era_name
            returning id
            """,
            (current_year, month_index, str(month_label), current_emperor, current_era, row),
        )
        time_point_id = int(cur.fetchone()["id"])
        time_points[row] = time_point_id

        for col, person_id in persons.items():
            cell = ws.cell(row, col)
            if cell.value in (None, ""):
                continue
            raw_text = str(cell.value).strip()
            event_type = classify_event_type(raw_text)
            if event_type == "tenure":
                continue
            source_cell = cell.coordinate
            cur = conn.execute(
                """
                insert into appointment_events (
                    person_id, time_point_id, event_type, raw_text, source_cell,
                    is_event_text, is_tenure_state, parse_confidence
                )
                values (?, ?, ?, ?, ?, ?, ?, ?)
                returning id
                """,
                (person_id, time_point_id, event_type, raw_text, source_cell, 1, 0, 0.6),
            )
            event_id = int(cur.fetchone()["id"])
            record_count += 1

            if cell.comment:
                conn.execute(
                    """
                    insert into annotations (event_id, source_cell, comment_text)
                    values (?, ?, ?)
                    """,
                    (event_id, source_cell, cell.comment.text),
                )
                comment_count += 1

    summary = {
        "source_file": str(source),
        "row_count": ws.max_row,
        "column_count": ws.max_column,
        "person_count": len(persons),
        "record_count": record_count,
        "comment_count": comment_count,
        "warning_count": warning_count,
    }
    conn.execute(
        """
        insert into import_audit (
            source_file, row_count, column_count, person_count,
            record_count, comment_count, warning_count
        )
        values (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            summary["source_file"],
            summary["row_count"],
            summary["column_count"],
            summary["person_count"],
            summary["record_count"],
            summary["comment_count"],
            summary["warning_count"],
        ),
    )
    conn.commit()
    conn.close()
    return summary
