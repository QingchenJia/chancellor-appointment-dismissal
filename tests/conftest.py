from pathlib import Path

import openpyxl
import pytest


@pytest.fixture
def temp_db_path(tmp_path: Path) -> Path:
    return tmp_path / "song_chancellors.db"


@pytest.fixture
def sample_workbook_path(tmp_path: Path) -> Path:
    workbook_path = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "宋代宰辅编年录"
    ws.append([])
    ws.append(["公元", "皇帝", "年号", "月份", "范质", "赵普（赵韩王）"])
    ws.append([960, "太祖赵匡胤", "建隆元年", "正月", None, None])
    ws.append([None, None, None, "二月", "本月乙亥日，自守司徒兼门下侍郎、同中书门下平章事，依前守司徒加兼侍中", None])
    ws.append([None, None, None, "三月", "侍中、同中书门下平章事", "本月甲申日，自右谏议大夫、枢密直学士升兵部侍郎，除枢密副使"])
    ws["F5"].comment = openpyxl.comments.Comment("《长编》日期有异文。", "tester")
    wb.create_sheet("说明")
    wb.save(workbook_path)
    return workbook_path
