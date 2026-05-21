import argparse
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.workbook, data_only=True)
    ws = wb.worksheets[0]
    person_columns = sum(1 for col in range(5, ws.max_column + 1) if ws.cell(2, col).value)
    records = 0
    comments = 0
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            if cell.column >= 5 and cell.value not in (None, ""):
                records += 1
            if cell.comment:
                comments += 1

    print(f"sheet: {ws.title}")
    print(f"rows: {ws.max_row}")
    print(f"columns: {ws.max_column}")
    print(f"person_columns: {person_columns}")
    print(f"nonempty_person_records: {records}")
    print(f"comments: {comments}")


if __name__ == "__main__":
    main()
